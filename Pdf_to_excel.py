# Take income statement pdf and put it into excel
# extract text from pdf
# clean up text
# put it in a database
# put it into excel spreadsheet

import PyPDF2
import sqlite3
import copy
import openpyxl
from openpyxl import Workbook


# extract string from pdf
pdfFileObj = open('Income_Statement.pdf','rb')
pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
pageObj = pdfReader.getPage(0)
stri = (pageObj.extractText())
print(stri)
stri =stri.translate({ord('\n'): None})

#get rid of unwanted characters in string
repl = ["&",":","(",")","/"," "]
for w in repl:
    stri = stri.replace(w,"")

text_arr = ['']
text_num = 0

letters ='abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ'
numbers ='0123456789.$,'

#Put string into list separated by line
i = 0

while((i<len(stri))and(stri[i])):
    if (i==len(stri)):
        break
    while(stri[i] not in numbers):
        text_arr[text_num]+=stri[i]
        i+=1
        if (i ==len(stri)):
            break
    if (i==len(stri)):
        break
    while(stri[i] in numbers):
        text_arr[text_num]+=stri[i]
        i+=1
        if (i==len(stri)):
            break
    text_arr.append('')
    text_num+=1
#print(text_arr)

num_arr = copy.deepcopy(text_arr)
cater_arr = copy.deepcopy(text_arr)


#delete letters in numarray
for letter in letters:
    k=0
    while(k<len(num_arr)):
        num_arr[k] = num_arr[k].replace(letter,"")
        k+=1


#delete numbers in letter array
for number in numbers:
    m=0
    while(m<len(cater_arr)):
        cater_arr[m]= cater_arr[m].replace(number,"")
        m+=1
    

#Part 2
#Now that you have catergories and numbers, put them into a database

conn = sqlite3.connect('excel.db')

cur = conn.cursor()

cur.executescript("""DROP TABLE IF EXISTS income;
            CREATE TABLE income(
            first text,
            amount text
            )""")

cur.execute("DELETE FROM income")

z=0

while (z<len(cater_arr)):
    cur.execute("INSERT INTO income VALUES (?,?)",(cater_arr[z],num_arr[z]))
    z+=1
conn.commit()
conn.close()

#Part 3
    #Put it into an excel spreadsheet
wb= Workbook()
ws = wb.active

#Put data from cater_arr and num_arr into spreadsheet
h = 1

while(h<len(cater_arr)):
    cellA = 'A'+ str(h)
    print(cellA)
    ws[cellA] = cater_arr[h-1]
    cellB = 'B'+ str(h)
    ws[cellB] = num_arr[h-1]

    h+=1

wb.save("sample.xlsx")










